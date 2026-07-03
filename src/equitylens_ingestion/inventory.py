"""Workbook structure inventory without transforming raw data."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import BinaryIO
from zipfile import ZipFile

import openpyxl
import xlrd

from equitylens_ingestion.models import JsonObject


def inspect_workbook(path: Path, expected_format: str) -> list[JsonObject]:
    """Return sheet names, order, visibility, and observed dimensions."""

    if expected_format == "xlsx":
        # A file object avoids openpyxl rejecting the required `.part` suffix.
        with path.open("rb") as handle:
            return _inspect_xlsx(handle)
    if expected_format == "xls":
        workbook = xlrd.open_workbook(path, on_demand=True)
        try:
            sheets: list[JsonObject] = []
            visibility_labels = {0: "visible", 1: "hidden", 2: "very_hidden"}
            for position, sheet_name in enumerate(workbook.sheet_names(), start=1):
                sheet = workbook.sheet_by_index(position - 1)
                sheets.append(
                    {
                        "position": position,
                        "name": sheet_name,
                        "visibility": visibility_labels.get(sheet.visibility, "unknown"),
                        "rows": sheet.nrows,
                        "columns": sheet.ncols,
                    }
                )
            return sheets
        finally:
            workbook.release_resources()
    if expected_format == "zip":
        sheets: list[JsonObject] = []
        with ZipFile(path) as archive:
            workbook_members = sorted(
                member for member in archive.namelist() if member.lower().endswith(".xlsx")
            )
            for member in workbook_members:
                sheets.extend(_inspect_xlsx(BytesIO(archive.read(member)), workbook_name=member))
        return sheets
    return []


def _inspect_xlsx(handle: BinaryIO, *, workbook_name: str | None = None) -> list[JsonObject]:
    workbook = openpyxl.load_workbook(handle, read_only=True, data_only=False)
    try:
        sheets: list[JsonObject] = []
        for position, sheet in enumerate(workbook.worksheets, start=1):
            rows, columns = _observed_dimensions(sheet)
            record: JsonObject = {
                "position": position,
                "name": sheet.title,
                "visibility": sheet.sheet_state,
                "rows": rows,
                "columns": columns,
            }
            if workbook_name is not None:
                record["workbook"] = workbook_name
            sheets.append(record)
        return sheets
    finally:
        workbook.close()


def _observed_dimensions(sheet: openpyxl.worksheet.worksheet.Worksheet) -> tuple[int, int]:
    rows = sheet.max_row or 0
    columns = sheet.max_column or 0
    if rows != 1 or columns != 1:
        return rows, columns

    # Some publisher workbooks incorrectly declare every sheet as A1. Scan only
    # that suspicious case so inventory reflects the actual cell range.
    sheet.reset_dimensions()
    observed_rows = 0
    observed_columns = 0
    for row in sheet.iter_rows():
        observed_rows += 1
        observed_columns = max(observed_columns, len(row))
    return observed_rows, observed_columns
